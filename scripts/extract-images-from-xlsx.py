"""
extract-images-from-xlsx.py

从 xlsx 文件抽取所有图片，输出到指定目录。
支持两种嵌入格式：
  1. WPS DISPIMG（xl/cellimages.xml + 单元格公式 _xlfn.DISPIMG）
  2. 标准 OOXML drawing（xl/drawings/drawingN.xml 锚点 + xl/media/*）—— 生意参谋导出格式
弹性设计：图片数量、行数、列数都不写死，实际有多少处理多少。

用法：
    python extract-images-from-xlsx.py <xlsx_path> <output_dir>

输出：stdout 最后一行是 JSON:
    {"images": [{"row": 2, "imageFile": "row02_xxx.jpeg", "name": "...", "brand": "...", "shop": "...", ...}]}
"""
import json
import os
import re
import sys
import zipfile

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

BRAND_KEYWORDS = [
    "喜临门", "麻大师", "源氏木语", "蓝盒子", "林氏", "金橡树", "亚朵",
    "全友", "雅兰", "海马", "梦百合", "芝华仕", "慕思", "simmons", "栖作",
]


def sanitize(s, maxlen=25):
    if not s:
        return "unknown"
    s = re.sub(r'[\\/:*?"<>|【】\[\]]', "", str(s))
    return s[:maxlen]


def detect_brand(shop):
    if not shop:
        return "unk"
    shop_lower = str(shop).lower()
    for kw in BRAND_KEYWORDS:
        if kw.lower() in shop_lower:
            return kw
    return "unk"


def find_column(headers, name_candidates):
    """按表头名找列索引，支持多个候选名"""
    for i, h in enumerate(headers):
        if not h:
            continue
        h_str = str(h).strip()
        for name in name_candidates:
            if name in h_str:
                return i + 1  # openpyxl 列从 1 开始
    return None


def parse_drawing_anchors(z):
    """解析标准 OOXML drawing 锚点（非 DISPIMG 的常规嵌入图片）。

    结构：xl/drawings/drawingN.xml 里每个 oneCellAnchor/twoCellAnchor 带
    <from><col>..</col><row>0基行号</row></from> 和 r:embed="rIdX"，
    对应的 rels 文件把 rIdX 映射到 ../media/imageN.jpeg。
    返回 {1基行号: media_target}，每行只保留第一张图。
    """
    names = set(z.namelist())
    row_to_target = {}
    drawing_files = sorted(
        n for n in names if re.fullmatch(r"xl/drawings/drawing\d+\.xml", n)
    )
    for df in drawing_files:
        rels_name = f"xl/drawings/_rels/{os.path.basename(df)}.rels"
        rid_to_target = {}
        if rels_name in names:
            rels_xml = z.read(rels_name).decode("utf-8", errors="ignore")
            # 属性顺序不固定（有的文件 Target 在前 Id 在后），逐个 Relationship 解析
            for m in re.finditer(r"<Relationship\b[^>]*/?>", rels_xml):
                tag = m.group(0)
                id_m = re.search(r'Id="(rId\d+)"', tag)
                target_m = re.search(r'Target="([^"]+)"', tag)
                if id_m and target_m:
                    rid_to_target[id_m.group(1)] = target_m.group(1)
        drawing_xml = z.read(df).decode("utf-8", errors="ignore")
        for m in re.finditer(
            r"<(?:oneCellAnchor|twoCellAnchor)\b.*?</(?:oneCellAnchor|twoCellAnchor)>",
            drawing_xml,
            re.DOTALL,
        ):
            block = m.group(0)
            row_m = re.search(r"<from>.*?<row>(\d+)</row>", block, re.DOTALL)
            embed_m = re.search(r'r:embed="(rId\d+)"', block)
            if not row_m or not embed_m:
                continue
            row_num = int(row_m.group(1)) + 1  # drawing XML 行号是 0 基
            target = rid_to_target.get(embed_m.group(1))
            if target and row_num not in row_to_target:
                row_to_target[row_num] = target
    return row_to_target


def main():
    if len(sys.argv) < 3:
        print(
            json.dumps({"error": "参数不足：需要 xlsx_path 和 output_dir"}),
            file=sys.stderr,
        )
        sys.exit(1)

    xlsx_path = sys.argv[1]
    output_dir = sys.argv[2]
    os.makedirs(output_dir, exist_ok=True)

    if not os.path.exists(xlsx_path):
        print(json.dumps({"error": f"xlsx 不存在: {xlsx_path}"}), file=sys.stderr)
        sys.exit(1)

    # 1. 读取 xlsx 内部 XML，建立 DISPIMG ID → media 文件 的映射
    try:
        with zipfile.ZipFile(xlsx_path) as z:
            sheet_files = [n for n in z.namelist() if n.startswith("xl/worksheets/sheet") and n.endswith(".xml")]
            sheet_xml = z.read(sheet_files[0]).decode("utf-8", errors="ignore") if sheet_files else ""
            cellimages_xml = z.read("xl/cellimages.xml").decode("utf-8", errors="ignore") if "xl/cellimages.xml" in z.namelist() else ""
            ci_rels = z.read("xl/_rels/cellimages.xml.rels").decode("utf-8", errors="ignore") if "xl/_rels/cellimages.xml.rels" in z.namelist() else ""
    except Exception as e:
        print(json.dumps({"error": f"读取 xlsx zip 失败: {e}"}), file=sys.stderr)
        sys.exit(1)

    id_to_embed = {}
    for m in re.finditer(r'name="(ID_[A-F0-9]+)".*?r:embed="(rId\d+)"', cellimages_xml, re.DOTALL):
        id_to_embed[m.group(1)] = m.group(2)

    rid_to_target = {}
    for m in re.finditer(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', ci_rels):
        rid_to_target[m.group(1)] = m.group(2)

    # 2. 扫描所有单元格中的 DISPIMG，记录 row 和列（不假设一定是 G 列）
    # 匹配 <c r="X999"> ... _xlfn.DISPIMG(&quot;ID_xxx&quot;
    row_col_to_id = {}
    for m in re.finditer(
        r'<c r="([A-Z]+)(\d+)"[^>]*>(?:<f[^>]*>|<v>)?_xlfn\.DISPIMG\(&quot;(ID_[A-F0-9]+)&quot;',
        sheet_xml,
    ):
        col_letter = m.group(1)
        row_num = int(m.group(2))
        disp_id = m.group(3)
        row_col_to_id[(row_num, col_letter)] = disp_id

    # 3. 加载 xlsx 读取商品上下文（按表头名智能查列）
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    col_name = find_column(headers, ["商品名称", "商品名", "name"])
    col_shop = find_column(headers, ["店铺名称", "店铺", "shop"])
    col_platform = find_column(headers, ["店铺类型", "平台", "platform"])
    col_price = find_column(headers, ["支付买家数", "价格", "price"])  # 原始表这列是价格带
    col_sales = find_column(headers, ["访客数", "月销", "sales"])
    col_keywords = find_column(headers, ["商品关键词", "关键词", "keywords"])
    col_ranking = find_column(headers, ["行业排名", "排名", "ranking"])
    col_date = find_column(headers, ["日期", "date"])

    # 4. 建立 行号 → media 目标 的统一映射（DISPIMG 优先，标准 drawing 兜底）
    row_to_target = {}
    for (row_num, col_letter), disp_id in sorted(row_col_to_id.items()):
        if row_num < 2:  # 跳过表头
            continue
        if row_num in row_to_target:
            continue  # 每行只取第一张主图
        rid = id_to_embed.get(disp_id)
        target = rid_to_target.get(rid) if rid else None
        if target:
            row_to_target[row_num] = target

    if not row_to_target:
        # 标准 OOXML drawing（生意参谋等导出的常规嵌入图）
        with zipfile.ZipFile(xlsx_path) as z:
            row_to_target = parse_drawing_anchors(z)

    # 5. 抽图
    images = []
    with zipfile.ZipFile(xlsx_path) as z:
        for row_num, target in sorted(row_to_target.items()):
            if row_num < 2:
                continue

            product_name = sanitize(ws.cell(row=row_num, column=col_name).value) if col_name else "unknown"
            shop = ws.cell(row=row_num, column=col_shop).value if col_shop else ""
            brand = detect_brand(shop)

            # 图片文件名
            data = None
            ext = ".jpeg"
            for candidate in [
                f"xl/{target.lstrip('./').lstrip('../')}",
                f"xl/media/{os.path.basename(target)}",
            ]:
                try:
                    data = z.read(candidate)
                    ext = os.path.splitext(candidate)[1] or ".jpeg"
                    break
                except KeyError:
                    continue

            if data is None:
                continue

            out_name = f"row{row_num:02d}_{brand}_{product_name}{ext}"
            out_path = os.path.join(output_dir, out_name)
            with open(out_path, "wb") as f:
                f.write(data)

            entry = {
                "row": row_num,
                "imageFile": out_name,
                "name": ws.cell(row=row_num, column=col_name).value if col_name else None,
                "brand": brand,
                "shop": shop,
                "platform": ws.cell(row=row_num, column=col_platform).value if col_platform else None,
                "price": ws.cell(row=row_num, column=col_price).value if col_price else None,
                "sales": ws.cell(row=row_num, column=col_sales).value if col_sales else None,
                "keywords": ws.cell(row=row_num, column=col_keywords).value if col_keywords else None,
                "ranking": ws.cell(row=row_num, column=col_ranking).value if col_ranking else None,
                "date": ws.cell(row=row_num, column=col_date).value if col_date else None,
            }
            # 转换日期对象为字符串
            if hasattr(entry["date"], "isoformat"):
                entry["date"] = entry["date"].isoformat()
            images.append(entry)

    # 5. 输出 JSON（最后一行）
    print(json.dumps({"images": images}, ensure_ascii=False))


if __name__ == "__main__":
    main()
