"""
read-ranking-rows.py

读取生意参谋排行 Excel 的商品行 + 图片链接（不抽取嵌入图片）。
同时做格式校验：缺少必需列时输出结构化错误。

用法：
    python read-ranking-rows.py <xlsx_path>

输出：stdout 最后一行 JSON:
    成功: {"ok": true, "period": "2026-08-25", "rows": [{"row": 2, "itemId": "...", "name": "...", "shop": "...", "imageUrl": "...", ...}]}
    失败: {"ok": false, "error": "缺少必需列: 商品图片链接", "missing": ["商品图片链接"], "headers": [...]}
"""
import json
import re
import sys

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

# 必需列（表头名候选）；图片链接和嵌入图片二选一
REQUIRED = {
    "name": ["商品名称", "商品名"],
    "itemId": ["商品ID"],
    "shop": ["店铺名称", "店铺"],
}
OPTIONAL = {
    "rank": ["行业排名", "排名"],
    "shopType": ["店铺类型", "平台"],
    "buyers": ["支付买家数"],
    "keywords": ["商品关键词", "关键词"],
    "date": ["日期"],
    "imageUrl": ["商品图片链接", "图片链接"],
    "imageEmbed": ["商品图片"],  # 嵌入图列（无链接时的存在性提示用）
}


def find_col(headers, candidates, prefer_numeric_row=None):
    """按表头名找列；同名多列时优先首个数据单元格为数字的列"""
    idxs = [i for i, h in enumerate(headers) if h and str(h).strip() in candidates]
    if not idxs:
        return None
    if len(idxs) > 1 and prefer_numeric_row:
        for i in idxs:
            v = prefer_numeric_row[i] if i < len(prefer_numeric_row) else None
            if v and str(v).strip().isdigit():
                return i
    return idxs[0]


def main():
    if len(sys.argv) < 2:
        print(json.dumps({"ok": False, "error": "参数不足"}, ensure_ascii=False))
        sys.exit(1)

    try:
        wb = openpyxl.load_workbook(sys.argv[1], data_only=True)
    except Exception as e:
        print(json.dumps({"ok": False, "error": f"文件无法作为 xlsx 打开: {e}"}, ensure_ascii=False))
        sys.exit(0)

    ws = wb.worksheets[0]
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    first_row = [str(c).strip() if c is not None else "" for c in next(ws.iter_rows(min_row=2, max_row=2, values_only=True))] if ws.max_row >= 2 else []

    cols = {}
    missing = []
    for key, candidates in REQUIRED.items():
        idx = find_col(headers, candidates, prefer_numeric_row=first_row if key == "itemId" else None)
        if idx is None:
            missing.append(candidates[0])
        cols[key] = idx
    for key, candidates in OPTIONAL.items():
        cols[key] = find_col(headers, candidates)

    if missing:
        print(json.dumps({
            "ok": False,
            "error": f"缺少必需列: {'、'.join(missing)}",
            "missing": missing,
            "headers": [h for h in headers if h],
        }, ensure_ascii=False))
        sys.exit(0)

    if cols["imageUrl"] is None and cols["imageEmbed"] is None:
        print(json.dumps({
            "ok": False,
            "error": "缺少图片列：需要「商品图片链接」或「商品图片」",
            "missing": ["商品图片链接/商品图片"],
            "headers": [h for h in headers if h],
        }, ensure_ascii=False))
        sys.exit(0)

    def cell(row, idx):
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    rows = []
    period = ""
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        item_id = str(cell(row, cols["itemId"]) or "").strip()
        if not item_id.isdigit():
            continue
        image_url = str(cell(row, cols["imageUrl"]) or "").strip() if cols["imageUrl"] is not None else ""
        if image_url.startswith("//"):
            image_url = "https:" + image_url
        date_val = cell(row, cols["date"]) if cols["date"] is not None else None
        if not period and date_val:
            period = str(date_val).strip()
        rank_val = cell(row, cols["rank"]) if cols["rank"] is not None else None
        rows.append({
            "row": r_idx,
            "itemId": item_id,
            "name": str(cell(row, cols["name"]) or "").strip(),
            "shop": str(cell(row, cols["shop"]) or "").strip(),
            "platform": str(cell(row, cols["shopType"]) or "").strip() if cols["shopType"] is not None else "",
            "ranking": int(rank_val) if rank_val is not None and str(rank_val).isdigit() else None,
            "date": str(date_val).strip() if date_val else "",
            "price": "",
            "sales": str(cell(row, cols["buyers"]) or "") if cols["buyers"] is not None else "",
            "keywords": str(cell(row, cols["keywords"]) or "").strip() if cols["keywords"] is not None else "",
            "imageUrl": image_url if image_url.startswith("http") else "",
        })

    if not rows:
        print(json.dumps({"ok": False, "error": "未读到任何有效商品行（商品ID 列为空或非数字）", "headers": [h for h in headers if h]}, ensure_ascii=False))
        sys.exit(0)

    print(json.dumps({"ok": True, "period": period, "rows": rows}, ensure_ascii=False))


if __name__ == "__main__":
    main()
